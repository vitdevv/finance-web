from fastapi import FastAPI, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from pydantic import BaseModel
import sqlite3, hashlib, os, tempfile
from datetime import datetime, timedelta
import requests as http_req
import jwt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SECRET = os.environ.get("JWT_SECRET", "finance_secret_key_change_in_production")
DB_PATH = os.environ.get("DB_PATH", os.path.join(os.path.dirname(os.path.abspath(__file__)), "finance.db"))
ALGORITHM = "HS256"

_raw_origins = os.environ.get("ALLOWED_ORIGINS", "http://localhost:5173,http://localhost:3000")
ALLOWED_ORIGINS = [o.strip() for o in _raw_origins.split(",")]

app = FastAPI(title="Finance PRO API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
security = HTTPBearer()

# ── Database ─────────────────────────────────────────────────────────────────

def get_conn():
    return sqlite3.connect(DB_PATH)

def init_db():
    with get_conn() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL, created_at TEXT DEFAULT (datetime('now')))""")
        conn.execute("""CREATE TABLE IF NOT EXISTS calculations (
            id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now')), usd_received REAL, usd_rate REAL,
            brl_extra REAL, prolabore REAL, regime TEXT, das REAL, inss REAL, irpf REAL,
            net_profit REAL, FOREIGN KEY (user_id) REFERENCES users(id))""")
        conn.execute("""CREATE TABLE IF NOT EXISTS assets (
            id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now')), type TEXT, name TEXT, amount REAL,
            FOREIGN KEY (user_id) REFERENCES users(id))""")
        conn.execute("""CREATE TABLE IF NOT EXISTS closed_months (
            id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL, label TEXT,
            closed_at TEXT DEFAULT (datetime('now')), net_profit REAL,
            total_deductions REAL, balance REAL, FOREIGN KEY (user_id) REFERENCES users(id))""")
        for sql in [
            "ALTER TABLE assets ADD COLUMN closed_month_id INTEGER DEFAULT NULL",
            "ALTER TABLE assets ADD COLUMN orig_amount REAL",
            "ALTER TABLE assets ADD COLUMN orig_currency TEXT DEFAULT 'BRL'",
            "ALTER TABLE assets ADD COLUMN rate_used REAL DEFAULT 1.0",
        ]:
            try:
                conn.execute(sql)
            except Exception:
                pass
        conn.execute("UPDATE assets SET orig_amount=amount, orig_currency='BRL', rate_used=1.0 WHERE orig_amount IS NULL")

init_db()

# ── Auth helpers ──────────────────────────────────────────────────────────────

def hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def make_token(uid: int, username: str) -> str:
    payload = {"user_id": uid, "username": username,
                "exp": datetime.utcnow() + timedelta(days=7)}
    return jwt.encode(payload, SECRET, algorithm=ALGORITHM)

def current_user(creds: HTTPAuthorizationCredentials = Depends(security)):
    try:
        p = jwt.decode(creds.credentials, SECRET, algorithms=[ALGORITHM])
        return {"id": p["user_id"], "username": p["username"]}
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

# ── Tax calculation helper ────────────────────────────────────────────────────

def calc_taxes(usd_received, usd_rate, brl_extra, prolabore, mode):
    receita_usd = usd_received * usd_rate
    receita_total = receita_usd + brl_extra
    if receita_total == 0:
        raise ValueError("Total revenue is zero")
    if mode == "PF":
        return dict(regime="PF", receita_usd=receita_usd, receita_total=receita_total,
                    prolabore=0, das=0, inss=0, irpf=0, net_profit=receita_total)
    pro = prolabore
    fator_r = pro / receita_total if receita_total else 0
    aliquota, regime = (0.06, "Anexo III") if fator_r >= 0.28 else (0.155, "Anexo V")
    das = receita_total * aliquota
    inss = pro * 0.11
    base_ir = pro - inss
    irpf = max(base_ir * 0.075 - 158, 0) if base_ir > 2000 else 0
    lucro = receita_total - das - inss - irpf
    return dict(regime=regime, receita_usd=receita_usd, receita_total=receita_total,
                prolabore=pro, das=das, inss=inss, irpf=irpf, net_profit=lucro)

# ── Pydantic models ───────────────────────────────────────────────────────────

class AuthReq(BaseModel):
    username: str
    password: str

class CalcReq(BaseModel):
    usd_received: float = 0
    usd_rate: float = 0
    brl_extra: float = 0
    prolabore: float = 0
    mode: str = "PJ"

class CalcEditReq(BaseModel):
    usd_received: float
    usd_rate: float
    brl_extra: float
    prolabore: float = 0
    date: str

class AssetReq(BaseModel):
    type: str
    name: str
    orig_amount: float
    orig_currency: str = "BRL"
    rate_used: float = 1.0
    amount: float

class AssetEditReq(BaseModel):
    name: str
    orig_amount: float
    rate_used: float
    amount: float

# ── Auth routes ───────────────────────────────────────────────────────────────

@app.post("/auth/register")
def register(b: AuthReq):
    try:
        with get_conn() as conn:
            conn.execute("INSERT INTO users (username, password) VALUES (?,?)",
                         (b.username.strip(), hash_pw(b.password)))
        return {"message": "Account created!"}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Username already exists.")

@app.post("/auth/login")
def login(b: AuthReq):
    with get_conn() as conn:
        row = conn.execute("SELECT id, username FROM users WHERE username=? AND password=?",
                           (b.username.strip(), hash_pw(b.password))).fetchone()
    if not row:
        raise HTTPException(status_code=401, detail="Invalid username or password.")
    return {"token": make_token(row[0], row[1]), "username": row[1]}

# ── Calculation routes ────────────────────────────────────────────────────────

@app.post("/calculations")
def calculate(b: CalcReq, u=Depends(current_user)):
    try:
        t = calc_taxes(b.usd_received, b.usd_rate, b.brl_extra, b.prolabore, b.mode)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    with get_conn() as conn:
        conn.execute("""INSERT INTO calculations
            (user_id, usd_received, usd_rate, brl_extra, prolabore, regime, das, inss, irpf, net_profit)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (u["id"], b.usd_received, b.usd_rate, b.brl_extra,
             t["prolabore"], t["regime"], t["das"], t["inss"], t["irpf"], t["net_profit"]))
    return t

@app.get("/calculations")
def get_calcs(u=Depends(current_user)):
    uid = u["id"]
    with get_conn() as conn:
        last_close = conn.execute(
            "SELECT closed_at FROM closed_months WHERE user_id=? ORDER BY closed_at DESC LIMIT 1", (uid,)
        ).fetchone()
        cutoff = last_close[0] if last_close else None
        rows = conn.execute("""SELECT id, created_at, usd_received, usd_rate, brl_extra, prolabore,
            regime, das, inss, irpf, net_profit FROM calculations
            WHERE user_id=? ORDER BY created_at DESC LIMIT 50""", (uid,)).fetchall()
    result = []
    for r in rows:
        cid, dt, usd_r, usd_rt, brl_e, pro, regime, das, inss, irpf, profit = r
        if cutoff and dt <= cutoff:
            continue
        result.append({"id": cid, "created_at": dt, "usd_received": usd_r, "usd_rate": usd_rt,
                        "brl_extra": brl_e, "prolabore": pro, "regime": regime, "das": das,
                        "inss": inss, "irpf": irpf, "net_profit": profit,
                        "revenue": usd_r * usd_rt + brl_e})
    return result

@app.delete("/calculations/{cid}")
def del_calc(cid: int, u=Depends(current_user)):
    with get_conn() as conn:
        conn.execute("DELETE FROM calculations WHERE id=? AND user_id=?", (cid, u["id"]))
    return {"message": "Deleted"}

@app.put("/calculations/{cid}")
def edit_calc(cid: int, b: CalcEditReq, u=Depends(current_user)):
    with get_conn() as conn:
        orig = conn.execute("SELECT regime FROM calculations WHERE id=? AND user_id=?",
                            (cid, u["id"])).fetchone()
    if not orig:
        raise HTTPException(status_code=404, detail="Not found")
    mode = "PF" if orig[0] == "PF" else "PJ"
    try:
        t = calc_taxes(b.usd_received, b.usd_rate, b.brl_extra, b.prolabore, mode)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    date_str = b.date if len(b.date) > 10 else b.date + " 00:00:00"
    with get_conn() as conn:
        conn.execute("""UPDATE calculations SET
            usd_received=?, usd_rate=?, brl_extra=?, prolabore=?, regime=?,
            das=?, inss=?, irpf=?, net_profit=?, created_at=?
            WHERE id=? AND user_id=?""",
            (b.usd_received, b.usd_rate, b.brl_extra, t["prolabore"], t["regime"],
             t["das"], t["inss"], t["irpf"], t["net_profit"], date_str, cid, u["id"]))
    return t

# ── Asset routes ──────────────────────────────────────────────────────────────

@app.post("/assets")
def add_asset(b: AssetReq, u=Depends(current_user)):
    with get_conn() as conn:
        conn.execute("""INSERT INTO assets
            (user_id, type, name, orig_amount, orig_currency, rate_used, amount)
            VALUES (?,?,?,?,?,?,?)""",
            (u["id"], b.type, b.name, b.orig_amount, b.orig_currency, b.rate_used, b.amount))
    return {"message": "Saved"}

@app.get("/assets")
def get_assets(u=Depends(current_user)):
    with get_conn() as conn:
        rows = conn.execute("""SELECT id, type, name,
            COALESCE(orig_amount, amount), COALESCE(orig_currency,'BRL'),
            COALESCE(rate_used,1.0), amount, created_at
            FROM assets WHERE user_id=? AND closed_month_id IS NULL
            ORDER BY created_at DESC""", (u["id"],)).fetchall()
    return [{"id": r[0], "type": r[1], "name": r[2], "orig_amount": r[3],
             "orig_currency": r[4], "rate_used": r[5], "amount": r[6], "created_at": r[7]}
            for r in rows]

@app.delete("/assets/{aid}")
def del_asset(aid: int, u=Depends(current_user)):
    with get_conn() as conn:
        conn.execute("DELETE FROM assets WHERE id=? AND user_id=?", (aid, u["id"]))
    return {"message": "Deleted"}

@app.put("/assets/{aid}")
def edit_asset(aid: int, b: AssetEditReq, u=Depends(current_user)):
    with get_conn() as conn:
        conn.execute("""UPDATE assets SET name=?, orig_amount=?, rate_used=?, amount=?, created_at=?
            WHERE id=? AND user_id=?""",
            (b.name, b.orig_amount, b.rate_used, b.amount,
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"), aid, u["id"]))
    return {"message": "Updated"}

@app.post("/assets/{aid}/refresh-rate")
def refresh_asset_rate(aid: int, u=Depends(current_user)):
    with get_conn() as conn:
        row = conn.execute("""SELECT orig_amount, COALESCE(orig_currency,'BRL'), name
            FROM assets WHERE id=? AND user_id=?""", (aid, u["id"])).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Not found")
    orig_amt, currency, name = row
    if currency == "BRL":
        return {"rate": 1.0, "amount": orig_amt}
    try:
        res = http_req.get(
            f"https://economia.awesomeapi.com.br/json/last/{currency}-BRL",
            timeout=5,
        )
        res.raise_for_status()
        data = res.json()
        rate = round(float(data[f"{currency}BRL"]["bid"]), 4)
    except Exception:
        try:
            res = http_req.get(
                f"https://api.exchangerate-api.com/v4/latest/{currency}",
                timeout=5,
            )
            res.raise_for_status()
            rate = round(res.json()["rates"]["BRL"], 4)
        except Exception:
            raise HTTPException(status_code=503, detail="Could not fetch exchange rate")
    new_brl = round(orig_amt * rate, 2)
    with get_conn() as conn:
        conn.execute("""UPDATE assets SET rate_used=?, amount=?, created_at=?
            WHERE id=? AND user_id=?""",
            (rate, new_brl, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), aid, u["id"]))
    return {"rate": rate, "amount": new_brl}

# ── Rate route ────────────────────────────────────────────────────────────────

@app.get("/rate/{currency}")
def get_rate(currency: str):
    if currency == "BRL":
        return {"rate": 1.0}
    # Primary: AwesomeAPI commercial rate (same base NOMAD uses)
    try:
        res = http_req.get(
            f"https://economia.awesomeapi.com.br/json/last/{currency}-BRL",
            timeout=5,
        )
        res.raise_for_status()
        data = res.json()
        rate = round(float(data[f"{currency}BRL"]["bid"]), 4)
        return {"rate": rate}
    except Exception:
        pass
    # Fallback: exchangerate-api
    try:
        res = http_req.get(
            f"https://api.exchangerate-api.com/v4/latest/{currency}",
            timeout=5,
        )
        res.raise_for_status()
        return {"rate": round(res.json()["rates"]["BRL"], 4)}
    except Exception:
        raise HTTPException(status_code=503, detail="Could not fetch exchange rate")

# ── Balance route ─────────────────────────────────────────────────────────────

@app.get("/balance")
def get_balance(u=Depends(current_user)):
    uid = u["id"]
    with get_conn() as conn:
        last_close = conn.execute(
            "SELECT closed_at FROM closed_months WHERE user_id=? ORDER BY closed_at DESC LIMIT 1", (uid,)
        ).fetchone()
        cutoff = last_close[0] if last_close else None
        calcs = conn.execute(
            "SELECT net_profit, created_at FROM calculations WHERE user_id=? ORDER BY created_at DESC LIMIT 50", (uid,)
        ).fetchall()
        open_assets = conn.execute(
            "SELECT amount FROM assets WHERE user_id=? AND closed_month_id IS NULL", (uid,)
        ).fetchall()
    filtered = [c for c in calcs if not cutoff or c[1] > cutoff]
    if not filtered:
        return {"balance": None, "net_profit": None, "deductions": 0}
    latest_profit = filtered[0][0]
    deductions = sum(r[0] for r in open_assets)
    return {"balance": latest_profit - deductions, "net_profit": latest_profit, "deductions": deductions}

# ── Month routes ──────────────────────────────────────────────────────────────

@app.post("/months/close")
def close_month(u=Depends(current_user)):
    uid = u["id"]
    with get_conn() as conn:
        last_close = conn.execute(
            "SELECT closed_at FROM closed_months WHERE user_id=? ORDER BY closed_at DESC LIMIT 1", (uid,)
        ).fetchone()
        cutoff = last_close[0] if last_close else None
        calcs = conn.execute(
            "SELECT net_profit, created_at FROM calculations WHERE user_id=? ORDER BY created_at DESC LIMIT 50", (uid,)
        ).fetchall()
        open_assets = conn.execute(
            "SELECT amount FROM assets WHERE user_id=? AND closed_month_id IS NULL", (uid,)
        ).fetchall()
    filtered = [c for c in calcs if not cutoff or c[1] > cutoff]
    if not filtered:
        raise HTTPException(status_code=400, detail="No calculation found for this month.")
    latest_profit = filtered[0][0]
    deductions = sum(r[0] for r in open_assets)
    balance = latest_profit - deductions
    now = datetime.now()
    label = now.strftime("%B %Y")
    cutoff_dt = now.strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO closed_months (user_id, label, net_profit, total_deductions, balance) VALUES (?,?,?,?,?)",
            (uid, label, latest_profit, deductions, balance))
        mid = cur.lastrowid
        conn.execute(
            "UPDATE assets SET closed_month_id=? WHERE user_id=? AND closed_month_id IS NULL AND created_at <= ?",
            (mid, uid, cutoff_dt))
    return {"message": f"{label} closed and archived!", "label": label, "balance": balance}

@app.get("/months")
def get_months(u=Depends(current_user)):
    with get_conn() as conn:
        rows = conn.execute("""SELECT id, label, closed_at, net_profit, total_deductions, balance
            FROM closed_months WHERE user_id=? ORDER BY closed_at DESC""", (u["id"],)).fetchall()
    return [{"id": r[0], "label": r[1], "closed_at": r[2], "net_profit": r[3],
             "total_deductions": r[4], "balance": r[5]} for r in rows]

@app.delete("/months/{mid}")
def del_month(mid: int, u=Depends(current_user)):
    with get_conn() as conn:
        conn.execute("UPDATE assets SET closed_month_id=NULL WHERE closed_month_id=?", (mid,))
        conn.execute("DELETE FROM closed_months WHERE id=? AND user_id=?", (mid, u["id"]))
    return {"message": "Deleted"}

# ── Export route ──────────────────────────────────────────────────────────────

@app.get("/export")
def export_excel(u=Depends(current_user)):
    uid = u["id"]
    with get_conn() as conn:
        calcs = conn.execute("""SELECT id, created_at, usd_received, usd_rate, brl_extra, prolabore,
            regime, das, inss, irpf, net_profit FROM calculations
            WHERE user_id=? ORDER BY created_at DESC""", (uid,)).fetchall()
        assets = conn.execute("""SELECT a.type, a.name, a.amount, a.created_at,
            COALESCE(cm.label, 'Open') FROM assets a
            LEFT JOIN closed_months cm ON a.closed_month_id = cm.id
            WHERE a.user_id=? ORDER BY a.created_at DESC""", (uid,)).fetchall()
        months = conn.execute("""SELECT label, closed_at, net_profit, total_deductions, balance
            FROM closed_months WHERE user_id=? ORDER BY closed_at DESC""", (uid,)).fetchall()
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = tmp.name
    tmp.close()
    _build_excel(u["username"], calcs, assets, months, path)
    filename = f"finance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return FileResponse(path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename=filename)

# ── Excel builder ─────────────────────────────────────────────────────────────

def _build_excel(username, calcs, assets, months, path):
    HDR_FILL   = PatternFill("solid", fgColor="1F3864")
    ACC_FILL   = PatternFill("solid", fgColor="2E75B6")
    ALT_FILL   = PatternFill("solid", fgColor="D9E1F2")
    POS_FILL   = PatternFill("solid", fgColor="E2EFDA")
    NEG_FILL   = PatternFill("solid", fgColor="FCE4D6")
    TITLE_FONT = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    HDR_FONT   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    BODY_FONT  = Font(name="Arial", size=10)
    BRL_FMT    = 'R$ #,##0.00;(R$ #,##0.00);"-"'
    PCT_FMT    = '0.0%'
    thin       = Side(style="thin", color="AAAAAA")
    bdr        = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.border = bdr; cell.alignment = center

    def style_data(ws, row, cols, alt=False, brl_cols=None):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            if alt: cell.fill = ALT_FILL
            cell.font = BODY_FONT; cell.border = bdr; cell.alignment = center
            if brl_cols and c in brl_cols: cell.number_format = BRL_FMT

    def pos_neg_fill(ws, row, col, value):
        ws.cell(row=row, column=col).fill = POS_FILL if (value or 0) >= 0 else NEG_FILL

    wb = openpyxl.Workbook()

    # Dashboard sheet
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    for col, w in zip(["A","B","C","D"], [30,22,22,22]):
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Finance Dashboard — {username}"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = HDR_FILL; ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="888888")
    ws["A2"].alignment = center

    latest_calc = calcs[0] if calcs else None
    open_assets = [a for a in assets if a[4] == "Open"]
    open_deducts = sum(a[2] for a in open_assets)

    ws["A4"] = "LATEST MONTH SNAPSHOT"
    ws["A4"].font = Font(name="Arial", size=11, bold=True, color="1F3864")
    kpis = [
        ("Net Profit (PJ)", latest_calc[10] if latest_calc else 0),
        ("Total Deductions", open_deducts),
        ("Current Balance", (latest_calc[10] - open_deducts) if latest_calc else 0),
        ("Months Archived", len(months)),
    ]
    for i, (label, val) in enumerate(kpis):
        col = i + 1
        lc = ws.cell(row=5, column=col, value=label)
        lc.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        lc.fill = ACC_FILL; lc.alignment = center; lc.border = bdr
        vc = ws.cell(row=6, column=col, value=val)
        vc.font = Font(name="Arial", size=13, bold=True, color="1F3864")
        vc.number_format = BRL_FMT if i < 3 else "0"
        vc.alignment = center; vc.border = bdr
        if i == 2:
            vc.fill = POS_FILL if val >= 0 else NEG_FILL
    ws.row_dimensions[6].height = 28

    total_rev = sum((c[2]*c[3] + c[4]) for c in calcs)
    total_das = sum(c[7] for c in calcs)
    total_inss = sum(c[8] for c in calcs)
    total_irpf = sum(c[9] for c in calcs)
    ws["A8"] = "CUMULATIVE TAX BURDEN"
    ws["A8"].font = Font(name="Arial", size=11, bold=True, color="1F3864")
    for ci, h in enumerate(["Category", "Total (R$)", "% of Revenue"], 1):
        c = ws.cell(row=9, column=ci, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = center; c.border = bdr
    for ri, (cat, val) in enumerate([("DAS", total_das), ("INSS", total_inss), ("IRPF", total_irpf)], 10):
        ws.cell(row=ri, column=1, value=cat).font = BODY_FONT
        ws.cell(row=ri, column=1).border = bdr
        v = ws.cell(row=ri, column=2, value=val)
        v.number_format = BRL_FMT; v.font = BODY_FONT; v.border = bdr; v.alignment = center
        pct = val / total_rev if total_rev else 0
        p = ws.cell(row=ri, column=3, value=pct)
        p.number_format = PCT_FMT; p.font = BODY_FONT; p.border = bdr; p.alignment = center

    # Calculations sheet
    ws2 = wb.create_sheet("Calculations")
    ws2.sheet_view.showGridLines = False
    for i, w in enumerate([18,12,14,10,12,12,10,12,12,12,14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.merge_cells("A1:K1")
    ws2["A1"] = "Calculation History"
    ws2["A1"].font = TITLE_FONT; ws2["A1"].fill = HDR_FILL
    ws2["A1"].alignment = center; ws2.row_dimensions[1].height = 24
    for ci, h in enumerate(["Date","Regime","USD Recv","Rate","USD→BRL","BRL Extra","Revenue","DAS","INSS","IRPF","Net Profit"], 1):
        ws2.cell(row=2, column=ci, value=h)
    style_header(ws2, 2, 11)
    brl_cols2 = {3,5,6,7,8,9,10,11}
    for ri, row in enumerate(calcs, 3):
        cid, dt, usd_r, usd_rt, brl_e, pro, regime, das, inss, irpf, profit = row
        revenue = usd_r * usd_rt + brl_e
        for ci, v in enumerate([dt[:16],regime,usd_r,usd_rt,usd_r*usd_rt,brl_e,revenue,das,inss,irpf,profit], 1):
            ws2.cell(row=ri, column=ci, value=v)
        style_data(ws2, ri, 11, alt=(ri%2==0), brl_cols=brl_cols2)
        pos_neg_fill(ws2, ri, 11, profit)

    # Assets sheet
    ws3 = wb.create_sheet("Assets & Expenses")
    ws3.sheet_view.showGridLines = False
    for col, w in zip(["A","B","C","D","E"], [18,25,16,16,14]):
        ws3.column_dimensions[col].width = w
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "Assets & Expenses"
    ws3["A1"].font = TITLE_FONT; ws3["A1"].fill = HDR_FILL
    ws3["A1"].alignment = center; ws3.row_dimensions[1].height = 24
    for ci, h in enumerate(["Month","Type","Name","Amount","Date"], 1):
        ws3.cell(row=2, column=ci, value=h)
    style_header(ws3, 2, 5)
    type_colors = {"Expense":"FCE4D6","Investment":"D9E1F2","Savings":"E2EFDA","Asset":"FFF2CC"}
    for ri, row in enumerate(assets, 3):
        atype, aname, aamount, adt, month_label = row
        for ci, v in enumerate([month_label,atype,aname,aamount,adt[:10]], 1):
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.font = BODY_FONT; cell.border = bdr; cell.alignment = center
            if ci == 4: cell.number_format = BRL_FMT
            color = type_colors.get(atype)
            if color: cell.fill = PatternFill("solid", fgColor=color)

    # Archived months sheet
    ws4 = wb.create_sheet("Archived Months")
    ws4.sheet_view.showGridLines = False
    for col, w in zip(["A","B","C","D","E"], [20,18,18,18,16]):
        ws4.column_dimensions[col].width = w
    ws4.merge_cells("A1:E1")
    ws4["A1"] = "Archived Months"
    ws4["A1"].font = TITLE_FONT; ws4["A1"].fill = HDR_FILL
    ws4["A1"].alignment = center; ws4.row_dimensions[1].height = 24
    for ci, h in enumerate(["Month","Closed At","Net Profit","Deductions","Balance"], 1):
        ws4.cell(row=2, column=ci, value=h)
    style_header(ws4, 2, 5)
    for ri, row in enumerate(months, 3):
        label, closed_at, profit, deductions, balance = row
        for ci, v in enumerate([label,closed_at[:16],profit,deductions,balance], 1):
            cell = ws4.cell(row=ri, column=ci, value=v)
            cell.font = BODY_FONT; cell.border = bdr; cell.alignment = center
            if ci in (3,4,5): cell.number_format = BRL_FMT
        pos_neg_fill(ws4, ri, 5, balance)

    wb.save(path)
